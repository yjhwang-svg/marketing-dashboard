# -*- coding: utf-8 -*-
"""
토스 파이프라인 코어 — 타겟 시트 기준 1)이미지 2)에어브릿지 3)엑셀(공유 드라이브 업로드).
시크릿은 st.secrets / .streamlit/secrets.toml / 환경변수에서 읽는다 (코드에 하드코딩 없음).
"""
import os, json, urllib.request, urllib.parse, urllib.error, base64, time, re, datetime
from io import BytesIO
from pathlib import Path

# ── 시트/스코프 상수 (비밀 아님) ──────────────────────────────────
TARGET_ID   = '11GuRNtcK2dqc82SlP9DD4BXKnU1z6ad6_62EWLJxWKc'
TARGET_GID  = 1445054260
SOURCE_ID   = '1FDog38MW7faYyHC4hXXcHg2e055cDDuOoH-3t9oeA-A'
SOURCE_GID  = 1340025873
AB_CHANNEL  = 'toss_beta'
SCOPES      = 'https://www.googleapis.com/auth/spreadsheets https://www.googleapis.com/auth/drive'

_HERE = Path(__file__).resolve().parent

# 엑셀 템플릿 후보 경로 (배포 시 repo 루트, 로컬은 프로젝트 폴더)
_TEMPLATE_CANDIDATES = [
    _HERE / 'template_upload.xlsx',
    _HERE / 'project' / 'active' / 'toss-pipeline-dashboard' / 'template_upload.xlsx',
]


def _template_path():
    for p in _TEMPLATE_CANDIDATES:
        if p.exists():
            return str(p)
    raise FileNotFoundError('template_upload.xlsx 를 찾을 수 없습니다')


# ── 설정 로더 (st.secrets → secrets.toml → env → 로컬파일) ────────
_CONFIG = None

def get_config():
    global _CONFIG
    if _CONFIG is not None:
        return _CONFIG
    sa = imgbb = ab = drive = pw = None

    # 1) Streamlit secrets (스트림릿 실행 시 .streamlit/secrets.toml 자동 로드)
    try:
        import streamlit as st
        s = st.secrets
        if 'gcp_service_account' in s:
            sa = dict(s['gcp_service_account'])
        imgbb = s.get('imgbb_key'); ab = s.get('airbridge_token')
        drive = s.get('drive_folder_id'); pw = s.get('app_password')
    except Exception:
        pass

    # 2) secrets.toml 직접 파싱 (CLI 등 비-스트림릿 실행)
    if sa is None:
        try:
            import tomllib
            tp = _HERE / '.streamlit' / 'secrets.toml'
            if tp.exists():
                d = tomllib.load(open(tp, 'rb'))
                sa = sa or d.get('gcp_service_account')
                imgbb = imgbb or d.get('imgbb_key'); ab = ab or d.get('airbridge_token')
                drive = drive or d.get('drive_folder_id'); pw = pw or d.get('app_password')
        except Exception:
            pass

    # 3) 환경변수
    imgbb = imgbb or os.environ.get('IMGBB_KEY')
    ab    = ab    or os.environ.get('AIRBRIDGE_TOKEN')
    drive = drive or os.environ.get('DRIVE_FOLDER_ID')
    pw    = pw    or os.environ.get('APP_PASSWORD')
    if sa is None and os.environ.get('GCP_SA_JSON'):
        sa = json.loads(os.environ['GCP_SA_JSON'])

    # 4) 로컬 SA 파일 (개발용 최후 폴백)
    if sa is None:
        local = Path(r'C:\Users\MADUP\Downloads\meta-ads-setting-37e31a60c6b3.json')
        if local.exists():
            sa = json.loads(local.read_text(encoding='utf-8'))

    _CONFIG = {'sa': sa, 'imgbb': imgbb, 'ab': ab, 'drive': drive, 'password': pw}
    return _CONFIG


# 타겟 시트 컬럼 (0-based)
COL = {
    'J_url':9, 'K_start':10, 'L_end':11, 'M_name':12, 'N_disc':13, 'O_price':14,
    'P_chcode':15, 'Q_chname':16,
    'S_channel':18, 'T_link':19, 'U_start':20, 'V_end':21, 'W_traffic':22,
    'X_outlink':23, 'Y_name':24, 'Z_disc':25, 'AA_price':26, 'AB_img':27, 'AC_status':28,
}
REQUIRED_SAB = [
    ('S_channel','S(채널명)'), ('T_link','T(랜딩URL)'), ('U_start','U(시작날짜)'),
    ('V_end','V(종료날짜)'), ('W_traffic','W(목표트래픽)'), ('X_outlink','X(아웃링크여부)'),
    ('Y_name','Y(상품명)'), ('Z_disc','Z(할인율)'), ('AA_price','AA(판매가)'), ('AB_img','AB(이미지URL)'),
]


# ── 인증 ──────────────────────────────────────────────────────────
def _b64url(d):
    if isinstance(d, str): d = d.encode()
    return base64.urlsafe_b64encode(d).rstrip(b'=').decode()

def get_token():
    sa = get_config()['sa']
    if not sa:
        raise RuntimeError('서비스계정 자격증명을 찾을 수 없습니다 (secrets 설정 필요)')
    now = int(time.time())
    h = _b64url(json.dumps({'alg':'RS256','typ':'JWT'}))
    p = _b64url(json.dumps({'iss':sa['client_email'], 'scope':SCOPES,
        'aud':sa['token_uri'], 'iat':now, 'exp':now+3600}))
    from cryptography.hazmat.primitives import hashes, serialization
    from cryptography.hazmat.primitives.asymmetric import padding
    pk = serialization.load_pem_private_key(sa['private_key'].encode(), password=None)
    sig = pk.sign('{}.{}'.format(h, p).encode(), padding.PKCS1v15(), hashes.SHA256())
    jwt = '{}.{}.{}'.format(h, p, _b64url(sig))
    data = urllib.parse.urlencode({
        'grant_type':'urn:ietf:params:oauth:grant-type:jwt-bearer', 'assertion':jwt}).encode()
    with urllib.request.urlopen(urllib.request.Request(sa['token_uri'], data=data), timeout=15) as r:
        return json.loads(r.read())['access_token']


def get_tab(token, sid, gid):
    url = 'https://sheets.googleapis.com/v4/spreadsheets/{}?fields=sheets.properties'.format(sid)
    req = urllib.request.Request(url, headers={'Authorization':'Bearer '+token})
    with urllib.request.urlopen(req, timeout=15) as r:
        meta = json.loads(r.read())
    return next(s['properties']['title'] for s in meta['sheets']
                if s['properties']['sheetId'] == gid)

def sheets_get(token, sid, rng):
    url = 'https://sheets.googleapis.com/v4/spreadsheets/{}/values/{}'.format(sid, urllib.parse.quote(rng))
    req = urllib.request.Request(url, headers={'Authorization':'Bearer '+token})
    with urllib.request.urlopen(req, timeout=15) as r:
        return json.loads(r.read()).get('values', [])

def sheets_write_cell(token, sid, rng, value):
    url = 'https://sheets.googleapis.com/v4/spreadsheets/{}/values/{}?valueInputOption=RAW'.format(
        sid, urllib.parse.quote(rng))
    body = json.dumps({'values':[[value]]}).encode('utf-8')
    req = urllib.request.Request(url, data=body, method='PUT',
        headers={'Authorization':'Bearer '+token, 'Content-Type':'application/json; charset=utf-8'})
    with urllib.request.urlopen(req, timeout=15) as r:
        return json.loads(r.read())


# ── 구글 드라이브 업로드 (공유 드라이브) ──────────────────────────
def drive_upload(token, name, data_bytes, folder_id, mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'):
    """공유 드라이브 폴더에 파일 업로드. 같은 이름 있으면 덮어쓰기. webViewLink 반환."""
    # 같은 이름 기존 파일 검색 → 있으면 업데이트
    q = "name='{}' and trashed=false and '{}' in parents".format(name.replace("'", "\\'"), folder_id)
    surl = ('https://www.googleapis.com/drive/v3/files?q={}&supportsAllDrives=true'
            '&includeItemsFromAllDrives=true&fields=files(id)').format(urllib.parse.quote(q))
    existing = None
    try:
        with urllib.request.urlopen(urllib.request.Request(surl, headers={'Authorization':'Bearer '+token}), timeout=15) as r:
            fs = json.loads(r.read()).get('files', [])
            if fs:
                existing = fs[0]['id']
    except Exception:
        pass

    boundary = '===tosspipeline==='
    if existing:
        meta = {}
        up = 'https://www.googleapis.com/upload/drive/v3/files/{}?uploadType=multipart&supportsAllDrives=true&fields=id,webViewLink'.format(existing)
        method = 'PATCH'
    else:
        meta = {'name': name, 'parents': [folder_id]}
        up = 'https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart&supportsAllDrives=true&fields=id,webViewLink'
        method = 'POST'

    pre = ('--{b}\r\nContent-Type: application/json; charset=UTF-8\r\n\r\n{m}\r\n'
           '--{b}\r\nContent-Type: {mime}\r\n\r\n').format(b=boundary, m=json.dumps(meta), mime=mime).encode()
    post = ('\r\n--{b}--\r\n').format(b=boundary).encode()
    body = pre + data_bytes + post
    req = urllib.request.Request(up, data=body, method=method,
        headers={'Authorization':'Bearer '+token,
                 'Content-Type':'multipart/related; boundary={}'.format(boundary)})
    with urllib.request.urlopen(req, timeout=40) as r:
        return json.loads(r.read()).get('webViewLink', '')


# ── 드라이브 폴더 해석/생성 ───────────────────────────────────────
def extract_folder_id(spec):
    """폴더 URL 또는 ID 문자열에서 ID 추출 (없으면 None)"""
    s = (spec or '').strip()
    m = re.search(r'/folders/([A-Za-z0-9_-]+)', s) or re.search(r'[?&]id=([A-Za-z0-9_-]+)', s)
    if m:
        return m.group(1)
    if re.fullmatch(r'[A-Za-z0-9_-]{15,}', s):
        return s
    return None

def _drive_get(token, fid, fields='id,name,driveId,mimeType'):
    url = 'https://www.googleapis.com/drive/v3/files/{}?supportsAllDrives=true&fields={}'.format(fid, fields)
    with urllib.request.urlopen(urllib.request.Request(url, headers={'Authorization':'Bearer '+token}), timeout=15) as r:
        return json.loads(r.read())

def _find_child_folder(token, drive_id, parent, name):
    q = ("name='{}' and '{}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
         ).format(name.replace("'", "\\'"), parent)
    url = ('https://www.googleapis.com/drive/v3/files?q={}&corpora=drive&driveId={}'
           '&includeItemsFromAllDrives=true&supportsAllDrives=true&fields=files(id,name)').format(
        urllib.parse.quote(q), drive_id)
    with urllib.request.urlopen(urllib.request.Request(url, headers={'Authorization':'Bearer '+token}), timeout=15) as r:
        fs = json.loads(r.read()).get('files', [])
    return fs[0]['id'] if fs else None

def _create_folder(token, parent, name):
    body = json.dumps({'name': name, 'mimeType': 'application/vnd.google-apps.folder', 'parents': [parent]}).encode()
    url = 'https://www.googleapis.com/drive/v3/files?supportsAllDrives=true&fields=id'
    req = urllib.request.Request(url, data=body, method='POST',
        headers={'Authorization':'Bearer '+token, 'Content-Type':'application/json'})
    with urllib.request.urlopen(req, timeout=15) as r:
        return json.loads(r.read())['id']

def resolve_drive_folder(token, drive_root, spec):
    """저장 폴더 결정. spec이:
    - 비어있으면 → 공유 드라이브 최상위(drive_root)
    - 폴더 URL/ID → 그 폴더 (공유 드라이브 소속인지 확인)
    - 폴더명/경로(a/b) → drive_root 아래에서 찾고 없으면 생성
    반환: {'ok':bool, 'id':folder_id, 'name':표시명, 'msg':err}
    """
    spec = (spec or '').strip()
    if not spec:
        return {'ok': True, 'id': drive_root, 'name': '(공유 드라이브 최상위)'}

    fid = extract_folder_id(spec)
    if fid:
        try:
            meta = _drive_get(token, fid)
        except urllib.error.HTTPError as e:
            return {'ok': False, 'msg': '폴더 접근 불가({}). 공유 드라이브 내 폴더인지/서비스계정 공유됐는지 확인하세요.'.format(e.code)}
        if meta.get('mimeType') != 'application/vnd.google-apps.folder':
            return {'ok': False, 'msg': '폴더가 아닙니다 (파일 링크인 듯).'}
        if not meta.get('driveId'):
            return {'ok': False, 'msg': '개인 드라이브 폴더입니다. 서비스계정은 공유 드라이브에만 저장 가능합니다.'}
        return {'ok': True, 'id': meta['id'], 'name': meta.get('name', fid)}

    # 경로/이름으로 처리 (없으면 생성)
    try:
        parent = drive_root
        segs = [s.strip() for s in re.split(r'[\\/]+', spec) if s.strip()]
        for seg in segs:
            child = _find_child_folder(token, drive_root, parent, seg)
            if not child:
                child = _create_folder(token, parent, seg)
            parent = child
        return {'ok': True, 'id': parent, 'name': '/'.join(segs)}
    except urllib.error.HTTPError as e:
        return {'ok': False, 'msg': '폴더 생성/조회 실패({})'.format(e.code)}
    except Exception as e:
        return {'ok': False, 'msg': '폴더 처리 오류: {}'.format(str(e)[:100])}


# ── 유틸 ──────────────────────────────────────────────────────────
def cell(row, key):
    i = COL[key]
    v = row[i] if i < len(row) else ''
    return (v or '').strip() if isinstance(v, str) else (v if v is not None else '')

def lo_id(url):
    m = re.search(r'(L[OE]\d+)', url or '')
    return m.group(1) if m else ''

def iso_to_yymmdd(s):
    m = re.match(r'(\d{4})-(\d{2})-(\d{2})', (s or '').strip())
    return (m.group(1)[2:] + m.group(2) + m.group(3)) if m else ''

def to_dt(s):
    try:
        return datetime.datetime.strptime((s or '').strip(), '%Y-%m-%d')
    except Exception:
        return s

def to_int(s):
    s = str(s).replace(',', '').strip()
    try:
        return int(s)
    except Exception:
        return s if s else 0

def load_target(token, tab):
    return sheets_get(token, TARGET_ID, "'{}'!A1:AC400".format(tab))

def load_source(token, stab):
    return sheets_get(token, SOURCE_ID, "'{}'!A1:N400".format(stab))


# ── 1) 이미지 링크 ────────────────────────────────────────────────
def get_lotteon_image_url(product_url):
    headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
    with urllib.request.urlopen(urllib.request.Request(product_url, headers=headers), timeout=15) as resp:
        html = resp.read().decode('utf-8', errors='replace')
    m = re.search(r'(https://contents\.lotteon\.com/itemimage/[^"\']+bndl_img/[^"\']+_1\.png)', html)
    if m:
        return m.group(1).split('/dims/')[0]
    m = re.search(r'(https://contents\.lotteon\.com/itemimage/[^"\']+_1\.(png|jpg))', html)
    if m:
        return m.group(1).split('/dims/')[0]
    return None

def image_to_imgbb(img_url, name):
    from PIL import Image
    with urllib.request.urlopen(urllib.request.Request(img_url, headers={'User-Agent':'Mozilla/5.0'}), timeout=15) as resp:
        data = resp.read()
    img = Image.open(BytesIO(data)).resize((308, 308), Image.LANCZOS)
    ts = int(time.time()) % 256
    try:
        if img.mode != 'RGBA':
            img = img.convert('RGBA')
        px = img.load(); r,g,b,a = px[0,0]; px[0,0] = (r,g,b,(a+ts)%256)
    except Exception:
        pass
    buf = BytesIO(); img.save(buf, 'PNG')
    img_b64 = base64.b64encode(buf.getvalue()).decode('ascii')
    key = get_config()['imgbb']
    post = urllib.parse.urlencode({'key':key, 'image':img_b64, 'name':name}).encode()
    with urllib.request.urlopen(urllib.request.Request('https://api.imgbb.com/1/upload', data=post), timeout=30) as resp:
        return json.loads(resp.read().decode('utf-8', errors='replace'))['data']['url']

def process_image(token, tab, row_num, row):
    j = cell(row, 'J_url')
    if 'lotteon.com' not in j:
        return {'ok': False, 'msg': 'J열에 롯데온 상품URL 없음'}
    try:
        img_url = get_lotteon_image_url(j.replace('\xa0', '').strip())
        if not img_url:
            return {'ok': False, 'msg': '상품페이지에서 이미지 URL 못 찾음'}
        name = '{}-308x308'.format(lo_id(j) or 'product_{}'.format(row_num))
        hosted = image_to_imgbb(img_url, name)
        sheets_write_cell(token, TARGET_ID, "'{}'!AB{}".format(tab, row_num), hosted)
        return {'ok': True, 'msg': '이미지 업로드 완료', 'url': hosted}
    except Exception as e:
        return {'ok': False, 'msg': '에러: {}'.format(str(e)[:120])}


# ── 2) 에어브릿지 링크 ────────────────────────────────────────────
def resolve_source_row(source_rows, prod_id, start_yymmdd, end_yymmdd):
    """소스 시트에서 상품ID+날짜로 매칭된 행 전체 반환 (없으면 None)"""
    for r in source_rows:
        g = (r[6] if len(r) > 6 else '').strip()
        h = (r[7] if len(r) > 7 else '').strip()
        if not g:
            continue
        parts = g.split('_')
        if (lo_id(h) == prod_id and len(parts) >= 2
                and parts[0] == start_yymmdd and parts[1] == end_yymmdd):
            return r
    for r in source_rows:
        g = (r[6] if len(r) > 6 else '').strip()
        parts = g.split('_')
        if len(parts) >= 2 and parts[0] == start_yymmdd and parts[1] == end_yymmdd:
            h = (r[7] if len(r) > 7 else '').strip()
            if prod_id and prod_id in h:
                return r
    return None

def airbridge_create(campaign, deeplink, fallback):
    """deeplink=소스 N열, fallback=소스 M열 을 그대로 사용 (트래킹 파라미터 포함 버전)."""
    payload = {
        'channel': AB_CHANNEL,
        'campaignParams': {'campaign': campaign},
        'deeplinkUrl': deeplink,
        'fallbackPaths': {'android': fallback, 'ios': fallback, 'desktop': fallback},
    }
    token = get_config()['ab']
    req = urllib.request.Request('https://api.airbridge.io/v1/tracking-links',
        data=json.dumps(payload).encode(), method='POST',
        headers={'Content-Type':'application/json', 'Authorization':'Bearer '+token})
    with urllib.request.urlopen(req, timeout=20) as r:
        res = json.loads(r.read())
    tl = res['data']['trackingLink']
    return tl['link']['click'], tl.get('shortUrl', '')

# M/N이 아직 준비 안 된 행을 가리키는 placeholder
_PLACEHOLDER_MARKERS = ['채널코드 업데이트 필요', '채널코드', '업데이트 필요']

def process_airbridge(token, tab, row_num, row, source_rows):
    j = cell(row, 'J_url')
    pid = lo_id(j)
    start = iso_to_yymmdd(cell(row, 'K_start'))
    end = iso_to_yymmdd(cell(row, 'L_end'))
    if not (j and start and end):
        return {'ok': False, 'msg': 'J/K/L(상품URL·날짜) 누락'}
    srow = resolve_source_row(source_rows, pid, start, end)
    if not srow:
        return {'ok': False, 'msg': '소스시트에서 매칭 실패 (상품ID={} {}~{})'.format(pid, start, end)}

    campaign = (srow[6] if len(srow) > 6 else '').strip()    # G
    fallback = (srow[12] if len(srow) > 12 else '').strip()  # M열 (웹 폴백 URL)
    deeplink = (srow[13] if len(srow) > 13 else '').strip()  # N열 (딥링크)

    if not fallback or not deeplink:
        return {'ok': False, 'msg': '소스시트 M/N(폴백·딥링크)이 비어있음 (캠페인 {})'.format(campaign)}
    if any(mk in fallback or mk in deeplink for mk in _PLACEHOLDER_MARKERS):
        return {'ok': False,
                'msg': '소스시트 M/N에 "채널코드 업데이트 필요" placeholder가 남아있음 → 채널코드 업데이트 후 다시 실행 (캠페인 {})'.format(campaign)}

    try:
        long_link, short = airbridge_create(campaign, deeplink, fallback)
        sheets_write_cell(token, TARGET_ID, "'{}'!T{}".format(tab, row_num), long_link)
        return {'ok': True, 'msg': '링크 생성 완료 ({})'.format(campaign),
                'link': long_link, 'short': short, 'campaign': campaign}
    except urllib.error.HTTPError as e:
        return {'ok': False, 'msg': 'Airbridge API {}: {}'.format(e.code, e.read().decode('utf-8','replace')[:100])}
    except Exception as e:
        return {'ok': False, 'msg': '에러: {}'.format(str(e)[:120])}


# ── 3) 엑셀 파일 → 공유 드라이브 업로드 ───────────────────────────
def validate_sab(row):
    return [label for key, label in REQUIRED_SAB if cell(row, key) in (None, '', ' ')]

def campaign_from_link(link):
    m = re.search(r'campaign=([^&]+)', link or '')
    return m.group(1) if m else ''

def build_xlsx_bytes(row):
    """행 → 템플릿 2행 채운 xlsx 바이트"""
    from openpyxl import load_workbook
    wb = load_workbook(_template_path())
    ws = wb.active
    ws['A2'] = cell(row, 'S_channel')
    ws['B2'] = cell(row, 'T_link')
    ws['C2'] = to_dt(cell(row, 'U_start'))
    ws['D2'] = to_dt(cell(row, 'V_end'))
    ws['E2'] = to_int(cell(row, 'W_traffic'))
    ws['F2'] = cell(row, 'X_outlink')
    ws['G2'] = cell(row, 'Y_name')
    ws['H2'] = to_int(cell(row, 'Z_disc'))
    ws['I2'] = to_int(cell(row, 'AA_price'))
    ws['J2'] = cell(row, 'AB_img')
    buf = BytesIO(); wb.save(buf)
    return buf.getvalue()

def process_excel(token, tab, row_num, row, dest_folder_id=None):
    """S~AB 검증 → xlsx 생성 → 공유 드라이브(또는 지정 폴더) 업로드 → AC='파일생성완료'."""
    need = []
    if not cell(row, 'AB_img'):
        need.append('1번(이미지 링크)')
    if not cell(row, 'T_link'):
        need.append('2번(에어브릿지 링크)')
    if need:
        return {'ok': False, 'blocked': True, 'reason': 'prereq', 'need': need,
                'msg': '{}행: {} 가 먼저 실행되어야 합니다 (해당 셀 비어있음)'.format(row_num, ' · '.join(need))}
    missing = validate_sab(row)
    if missing:
        return {'ok': False, 'blocked': True, 'reason': 'cells', 'missing': missing,
                'msg': '{}행: {} 비어있음 → 채워주세요'.format(row_num, ', '.join(missing))}

    campaign = campaign_from_link(cell(row, 'T_link'))
    if not campaign:
        return {'ok': False, 'msg': 'T열 링크에서 campaign 값 추출 실패'}

    drive_id = dest_folder_id or get_config()['drive']
    try:
        data = build_xlsx_bytes(row)
    except Exception as e:
        return {'ok': False, 'msg': '엑셀 생성 실패: {}'.format(str(e)[:120])}

    link = ''
    if drive_id:
        try:
            link = drive_upload(token, '{}.xlsx'.format(campaign), data, drive_id)
        except urllib.error.HTTPError as e:
            return {'ok': False, 'msg': '드라이브 업로드 실패 {}: {}'.format(e.code, e.read().decode('utf-8','replace')[:120])}
        except Exception as e:
            return {'ok': False, 'msg': '드라이브 업로드 실패: {}'.format(str(e)[:120])}

    sheets_write_cell(token, TARGET_ID, "'{}'!AC{}".format(tab, row_num), '파일생성완료')
    return {'ok': True, 'msg': '파일 생성·업로드 완료', 'campaign': campaign, 'link': link, 'data': data}
